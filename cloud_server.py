# -*- coding: utf-8 -*-
"""
期货品种指数模型扫描器 v1.4 - 云端部署版 Web服务

部署到云端后，手机打开网址 -> 点"运行模型" -> 数据自动更新

支持平台：
  - Render (推荐，免费，最简单)
  - Hugging Face Spaces (免费)
  - Zeabur (免费)
  - 任何支持 Docker 的云服务器
"""

from __future__ import annotations

import json
import math
import os
import subprocess
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

APP_DIR = Path(__file__).parent
OUTPUT_DIR = APP_DIR / "output"
DATABASE_DIR = APP_DIR / "database"
CONFIG_FILE = APP_DIR / "config.json"

# 确保目录存在
OUTPUT_DIR.mkdir(exist_ok=True)
DATABASE_DIR.mkdir(exist_ok=True)

app = Flask(__name__, static_folder=None)
CORS(app)

# 全局运行状态
run_status = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "success": None,
    "output_file": None,
    "logs": [],
    "progress": "",
}


def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def find_latest_xlsx() -> Optional[Path]:
    if not OUTPUT_DIR.exists():
        return None
    files = sorted(OUTPUT_DIR.glob("期货模型交易看板_*.xlsx"), reverse=True)
    return files[0] if files else None


def read_sheet(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, header=None)
    if df.shape[0] >= 3:
        columns = df.iloc[1].astype(str).tolist()
        seen = set()
        final_cols = []
        for c in columns:
            if c in seen:
                final_cols.append(f"{c}_{len(seen)}")
            else:
                final_cols.append(c)
            seen.add(c)
        df = df.iloc[2:].reset_index(drop=True)
        df.columns = final_cols
    elif df.shape[0] >= 2:
        df.columns = df.iloc[0].astype(str).tolist()
        df = df.iloc[1:].reset_index(drop=True)
    return df


def df_to_json_records(df: pd.DataFrame) -> list:
    records = []
    for _, row in df.iterrows():
        record = {}
        for col in df.columns:
            val = row[col]
            if val is None:
                record[col] = None
            elif isinstance(val, float) and math.isnan(val):
                record[col] = None
            elif isinstance(val, (pd.Timestamp, datetime)):
                record[col] = val.strftime("%Y-%m-%d %H:%M:%S") if hasattr(val, "hour") and val.hour > 0 else val.strftime("%Y-%m-%d")
            elif isinstance(val, float):
                record[col] = round(val, 4)
            else:
                record[col] = str(val) if val is not None else None
        records.append(record)
    return records


def get_all_results(xlsx_path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    result = {"file_name": xlsx_path.name, "file_time": datetime.fromtimestamp(xlsx_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")}
    for sheet_name in sheets:
        try:
            df = read_sheet(xlsx_path, sheet_name)
            result[sheet_name] = df_to_json_records(df)
        except Exception as e:
            result[sheet_name] = {"error": str(e)}
    return result


def list_output_files():
    files = []
    if OUTPUT_DIR.exists():
        for f in sorted(OUTPUT_DIR.glob("期货模型交易看板_*.xlsx"), reverse=True):
            stat = f.stat()
            files.append({
                "name": f.name,
                "size": stat.st_size,
                "time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })
    return files


def read_latest_log():
    if not OUTPUT_DIR.exists():
        return ""
    logs = sorted(OUTPUT_DIR.glob("运行日志_*.txt"), reverse=True)
    if logs:
        return logs[0].read_text(encoding="utf-8")
    return ""


# ============ 路由 ============

@app.route("/")
def index():
    return send_from_directory(APP_DIR, "web_dashboard.html")


@app.route("/api/status")
def api_status():
    latest = find_latest_xlsx()
    status = {
        "run": {k: run_status[k] for k in ["running", "started_at", "finished_at", "success", "output_file", "progress"]},
        "latest_file": latest.name if latest else None,
        "latest_time": datetime.fromtimestamp(latest.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S") if latest else None,
        "config": load_config(),
    }
    return jsonify(status)


@app.route("/api/list-files")
def api_list_files():
    return jsonify({"files": list_output_files()})


@app.route("/api/data")
def api_data():
    file_index = request.args.get("index")
    file_name = request.args.get("file")

    if not file_index and not file_name:
        latest = find_latest_xlsx()
        if not latest:
            return jsonify({"error": "没有找到任何输出文件，请先运行模型"}), 404
        xlsx_path = latest
    elif file_index is not None:
        all_files = sorted(OUTPUT_DIR.glob("期货模型交易看板_*.xlsx"), reverse=True)
        idx = int(file_index)
        if 0 <= idx < len(all_files):
            xlsx_path = all_files[idx]
        else:
            return jsonify({"error": f"文件索引不存在: {idx}"}), 404
    else:
        from urllib.parse import unquote
        file_name = unquote(file_name)
        xlsx_path = OUTPUT_DIR / file_name
        if not xlsx_path.exists():
            return jsonify({"error": f"文件不存在: {file_name}"}), 404

    try:
        data = get_all_results(xlsx_path)
        data["log"] = read_latest_log()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/run", methods=["POST"])
def api_run():
    """触发模型计算。"""
    if run_status["running"]:
        return jsonify({"error": "模型正在运行中，请等待完成"}), 409

    mock = request.args.get("mock", "false").lower() == "true"

    def run_model():
        run_status["running"] = True
        run_status["started_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        run_status["finished_at"] = None
        run_status["success"] = None
        run_status["output_file"] = None
        run_status["logs"] = []
        run_status["progress"] = "正在启动模型计算..."

        try:
            cmd = [sys.executable, str(APP_DIR / "futures_scanner.py")]
            if mock:
                cmd.append("--mock")

            process = subprocess.Popen(
                cmd,
                cwd=str(APP_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                env={**os.environ, "PYTHONUTF8": "1", "PYTHONIOENCODING": "utf-8"},
            )

            for line in process.stdout:
                line = line.strip()
                if line:
                    run_status["logs"].append(line)
                    if "更新行情" in line and "%" in line:
                        run_status["progress"] = f"更新行情: {line}"
                    elif "计算因子" in line and "%" in line:
                        run_status["progress"] = f"计算因子与回测: {line}"
                    elif "输出 Excel" in line:
                        run_status["progress"] = line
                    elif "完成" in line:
                        run_status["progress"] = line

            process.wait()

            if process.returncode == 0:
                run_status["success"] = True
                run_status["progress"] = "运行成功！"
                latest = find_latest_xlsx()
                if latest:
                    run_status["output_file"] = latest.name
            else:
                run_status["success"] = False
                run_status["progress"] = f"运行失败（退出码 {process.returncode}）"

        except Exception as e:
            run_status["success"] = False
            run_status["progress"] = f"运行异常: {str(e)}"
            run_status["logs"].append(traceback.format_exc())

        finally:
            run_status["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            run_status["running"] = False

    thread = threading.Thread(target=run_model, daemon=True)
    thread.start()

    return jsonify({"message": "模型计算已启动", "mock": mock})


@app.route("/api/run-progress")
def api_run_progress():
    return jsonify({"run": run_status})


@app.route("/api/download/<filename>")
def api_download(filename):
    filepath = OUTPUT_DIR / filename
    if not filepath.exists():
        return jsonify({"error": "文件不存在"}), 404
    return send_from_directory(str(OUTPUT_DIR), filename, as_attachment=True)


@app.route("/web_dashboard.html")
def serve_dashboard():
    return send_from_directory(APP_DIR, "web_dashboard.html")


# ============ 入口 ============

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 60)
    print("期货模型交易看板 v1.4 - 云端版")
    print(f"端口: {port}")
    print("=" * 60)
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
