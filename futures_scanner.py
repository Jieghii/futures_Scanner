# -*- coding: utf-8 -*-
"""
期货品种指数/加权合约模型扫描器 v1.4

核心口径：
1) 使用品种连续/指数口径数据（如 RB0、LH0）做判断；
2) 信号在 t 日收盘后产生，验证 t+1 日开盘到收盘方向；
3) 13 个因子：删除“低买高卖”后，用户自定义 6 个 + 技术分析体系 7 个；
4) 自动更新本地数据库，输出 Excel 交易看板；
5) 增加数据新鲜度诊断，区分“接口尚未发布今日数据”和“更新失败”；
6) 增加近10日平均/最小价格振幅；
7) 盘中运行时尝试抓取实时快照，用于当前价、盘中高低点和盘中信号，盘后保持日线逻辑不变。

运行：
    python futures_scanner.py
模拟测试：
    python futures_scanner.py --mock
"""

from __future__ import annotations

import argparse
import json
import math
import os
import re
import sys
import traceback
import warnings
from dataclasses import dataclass
from datetime import datetime, time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore", category=FutureWarning)

try:
    from tqdm import tqdm
except Exception:  # pragma: no cover
    tqdm = lambda x, **kwargs: x


# -----------------------------
# 基础配置
# -----------------------------

FACTOR_NAMES = [
    "前高前低",
    "大阳大阴",
    "跳空缺口",
    "颈线趋势线",
    "整数关口",
    "黄金分割",
    "趋势方向",
    "支撑阻力",
    "成交量确认",
    "持仓量确认",
    "移动均线",
    "形态突破",
    "波动率变化",
]

PARAM_EXPLAIN = [
    ["数据口径", "品种指数/连续合约", "使用公开接口中的连续合约数据，如 RB0、LH0，作为品种指数/加权口径的近似。"],
    ["预测目标", "下一交易日开盘到收盘", "t 日收盘后出信号，用 t+1 日 close/open-1 验证。"],
    ["排序方式", "综合排序", "胜率、平均收益、盈亏比、最大回撤、交易次数加权。"],
    ["常规模型", "最低 20 次信号", "避免一年只出现极少次数的高胜率假象。"],
    ["低频模型", "最低 5 次信号", "单独输出高胜率低频机会。"],
    ["流动性过滤", "分位数过滤 + 数据时效", "剔除成交量、持仓量偏低或数据明显滞后的品种。"],
    ["数据新鲜度", "自动诊断", "若15:00刚收盘后运行，日频/结算数据可能尚未同步；数据状态sheet会提示是否等待重试。"],
    ["振幅字段", "近10日平均/最小价格振幅", "基于已完成日线的 high-low 计算，不使用未收盘的盘中快照，以免污染回测。"],
    ["盘中实时", "实时快照，不覆盖日线库", "交易时段内尝试通过 futures_zh_spot 抓取主力合约最新价、高低点、成交量、持仓；回测仍使用完整日线。"],
]

FACTOR_EXPLAIN = [
    ["前高前低", "前高/前低需经过后 5 日确认；突破前高偏多，跌破前低偏空，接近支撑/压力给反向提示。"],
    ["大阳大阴", "当日实体涨跌大于近 7 日平均波幅 2 倍触发；大阳偏多，大阴偏空，并参考大阳/大阴半分位。"],
    ["跳空缺口", "今日低点高于昨日高点为上跳空，偏多；今日高点低于昨日低点为下跳空，偏空。"],
    ["颈线趋势线", "近似使用已确认高低点连接成趋势线，向上突破下降压力线偏多，向下跌破上升支撑线偏空。"],
    ["整数关口", "根据价格级别自动识别整数关口，站上关口偏多，跌破关口偏空。"],
    ["黄金分割", "使用近 60 日高低点计算 0.382、0.5、0.618，向上突破偏多，向下跌破偏空。"],
    ["趋势方向", "MA20、MA60 与价格位置判断趋势方向。"],
    ["支撑阻力", "近 20 日区间突破高点偏多，跌破低点偏空。"],
    ["成交量确认", "价格上涨且成交量放大偏多；价格下跌且成交量放大偏空。"],
    ["持仓量确认", "价格与持仓同向增加代表资金推动，辅助确认趋势。"],
    ["移动均线", "MA5 与 MA20 金叉偏多，死叉偏空。"],
    ["形态突破", "低波动压缩后向上突破偏多，向下突破偏空。"],
    ["波动率变化", "ATR 或布林带宽度扩张且价格同向，说明行情进入可交易状态。"],
]


@dataclass
class BestModel:
    symbol: str
    name: str
    min_factors: int
    net_gap: int
    trades: int
    win_rate: float
    avg_return: float
    profit_factor: float
    max_drawdown: float
    composite_score: float
    direction_bias: str
    model_type: str


# -----------------------------
# 工具函数
# -----------------------------

def load_config(path: str = "config.json") -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def ensure_dirs(config: dict) -> None:
    Path(config["database_dir"]).mkdir(exist_ok=True)
    Path(config["output_dir"]).mkdir(exist_ok=True)


def safe_float(x, default=np.nan) -> float:
    try:
        if pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def normalize_daily_df(df: pd.DataFrame) -> pd.DataFrame:
    """兼容不同接口列名（新浪 + 东方财富）。"""
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.copy()
    rename_map = {
        # 新浪列名
        "日期": "date",
        "开盘价": "open",
        "最高价": "high",
        "最低价": "low",
        "收盘价": "close",
        "成交量": "volume",
        "持仓量": "hold",
        "动态结算价": "settle",
        # 东方财富列名
        "日期": "date",
        "开盘": "open",
        "最高": "high",
        "最低": "low",
        "收盘": "close",
        "成交量": "volume",
        "持仓量": "hold",
        "结算价": "settle",
    }
    df.rename(columns=rename_map, inplace=True)
    need = ["date", "open", "high", "low", "close", "volume", "hold", "settle"]
    for col in need:
        if col not in df.columns:
            if col == "settle" and "close" in df.columns:
                df[col] = df["close"]
            elif col in ["volume", "hold"]:
                df[col] = 0
            else:
                df[col] = np.nan
    df = df[need]
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    for col in ["open", "high", "low", "close", "volume", "hold", "settle"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df.dropna(subset=["date", "open", "high", "low", "close"])
    df = df.sort_values("date").drop_duplicates("date", keep="last")
    df = df[df["close"] > 0]
    df.reset_index(drop=True, inplace=True)
    return df


def symbol_base(symbol: str) -> str:
    return re.sub(r"0$", "", str(symbol).upper())


def clean_name(name: str) -> str:
    name = str(name)
    return name.replace("连续", "").replace("期货", "").strip()


def sign_to_text(x: int) -> str:
    if x > 0:
        return "多"
    if x < 0:
        return "空"
    return "中性"


def calc_max_drawdown(returns: pd.Series) -> float:
    if returns.empty:
        return 0.0
    equity = (1 + returns.fillna(0)).cumprod()
    peak = equity.cummax()
    dd = equity / peak - 1
    return float(dd.min())


def profit_factor(returns: pd.Series) -> float:
    gains = returns[returns > 0].sum()
    losses = -returns[returns < 0].sum()
    if losses <= 1e-12:
        return 9.99 if gains > 0 else 0.0
    return float(min(gains / losses, 9.99))


def score_model(win_rate: float, avg_return: float, pf: float, max_dd: float, trades: int, weights: dict) -> float:
    # 各项压缩到 0-100。收益为日内收益，0.5% 已经算较好，2% 封顶。
    win_score = max(0.0, min(100.0, win_rate * 100))
    avg_score = max(0.0, min(100.0, (avg_return + 0.002) / 0.012 * 100))
    pf_score = max(0.0, min(100.0, pf / 2.5 * 100))
    dd_score = max(0.0, min(100.0, 100 + max_dd * 500))  # -20% 约为 0 分
    trade_score = max(0.0, min(100.0, trades / 60 * 100))
    return float(
        weights.get("win_rate", 0.4) * win_score
        + weights.get("avg_return", 0.25) * avg_score
        + weights.get("profit_factor", 0.15) * pf_score
        + weights.get("drawdown", 0.1) * dd_score
        + weights.get("trades", 0.1) * trade_score
    )


def diagnose_data_freshness(latest_date: pd.Timestamp, run_ts: datetime) -> str:
    """判断日线数据是否真的更新到应有交易日。

    说明：程序使用的是日频数据，包含结算价。国内期货日盘 15:00 刚结束后，
    交易所和第三方行情源通常还需要生成/同步当日收盘、结算、成交量、持仓等日线字段。
    因此 15:02 看到最新日期仍为上一交易日，更多是“数据源尚未发布今日日线”，
    不应简单等同于程序更新失败。
    """
    if latest_date is None or pd.isna(latest_date):
        return "无有效日期"
    latest_d = pd.Timestamp(latest_date).date()
    today = run_ts.date()
    now_t = run_ts.time()

    if latest_d == today:
        return "已更新到运行日"
    if latest_d > today:
        return "日期异常：数据日期晚于运行日"

    # 仅按自然日做提示，不强行判断节假日。节假日/非交易日时，昨日或上个交易日是正常的。
    if run_ts.weekday() >= 5:
        return "周末/非交易日可能正常：最新数据停留在上一交易日"
    if now_t < time(15, 0):
        return "日盘尚未收盘：最新日线停留在上一交易日属正常"
    if time(15, 0) <= now_t < time(15, 45):
        return "15点刚收盘：日频/结算数据通常尚未同步，建议15:45后重试"
    if time(15, 45) <= now_t < time(18, 30):
        return "盘后同步窗口：若仍未更新，多半是第三方数据源延迟，建议稍后重试"
    return "运行日盘后仍未更新：可能是数据源延迟、接口异常、节假日或交易所尚未发布"


# -----------------------------
# 数据接口
# -----------------------------

def get_akshare():
    try:
        import akshare as ak  # type: ignore
        return ak
    except Exception as e:
        raise RuntimeError(
            "未能导入 akshare。请先运行：python -m pip install -r requirements.txt"
        ) from e


def fetch_symbol_list(config: dict, mock: bool = False) -> pd.DataFrame:
    if mock:
        return pd.DataFrame(
            [
                {"symbol": "RB0", "exchange": "shfe", "name": "螺纹钢连续"},
                {"symbol": "LH0", "exchange": "dce", "name": "生猪连续"},
                {"symbol": "I0", "exchange": "dce", "name": "铁矿石连续"},
                {"symbol": "SA0", "exchange": "czce", "name": "纯碱连续"},
                {"symbol": "CU0", "exchange": "shfe", "name": "沪铜连续"},
                {"symbol": "AU0", "exchange": "shfe", "name": "沪金连续"},
                {"symbol": "M0", "exchange": "dce", "name": "豆粕连续"},
                {"symbol": "CF0", "exchange": "czce", "name": "棉花连续"},
            ]
        )

    ak = get_akshare()
    errors = []
    # 第一优先：新浪品种列表
    try:
        df = ak.futures_display_main_sina()
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        if "symbol" in df.columns:
            if "exchange" not in df.columns:
                df["exchange"] = ""
            if "name" not in df.columns:
                df["name"] = df["symbol"]
            df["symbol"] = df["symbol"].astype(str).str.upper().str.strip()
            df["exchange"] = df["exchange"].astype(str).str.lower().str.strip()
            df["name"] = df["name"].astype(str).str.strip()
            excludes = set(x.upper() for x in config.get("exclude_symbols", []))
            if excludes:
                df = df[~df["symbol"].isin(excludes)]
            preferred = [x.upper() for x in config.get("preferred_symbols", [])]
            if preferred:
                df = df[df["symbol"].isin(preferred)]
            result = df[["symbol", "exchange", "name"]].drop_duplicates("symbol")
            if len(result) > 0:
                return result
    except Exception as e:
        errors.append(f"新浪品种列表失败: {e}")
    # 第二优先：东方财富品种列表（海外可用）
    try:
        df = ak.futures_zh_spot_sina()
        df = df.copy()
        df.columns = [str(c).strip() for c in df.columns]
        if "symbol" in df.columns:
            if "exchange" not in df.columns:
                df["exchange"] = ""
            if "name" not in df.columns:
                df["name"] = df["symbol"]
            df["symbol"] = df["symbol"].astype(str).str.upper().str.strip()
            df["exchange"] = df["exchange"].astype(str).str.lower().str.strip()
            df["name"] = df["name"].astype(str).str.strip()
            excludes = set(x.upper() for x in config.get("exclude_symbols", []))
            if excludes:
                df = df[~df["symbol"].isin(excludes)]
            preferred = [x.upper() for x in config.get("preferred_symbols", [])]
            if preferred:
                df = df[df["symbol"].isin(preferred)]
            result = df[["symbol", "exchange", "name"]].drop_duplicates("symbol")
            if len(result) > 0:
                return result
    except Exception as e:
        errors.append(f"东财品种列表失败: {e}")
    raise RuntimeError(f"获取品种列表失败: {'; '.join(errors)}")


def generate_mock_daily(symbol: str, n: int = 320) -> pd.DataFrame:
    rng = np.random.default_rng(abs(hash(symbol)) % (2**32))
    dates = pd.bdate_range(end=pd.Timestamp.today().normalize(), periods=n)
    base = {
        "RB0": 3500,
        "LH0": 16000,
        "I0": 800,
        "SA0": 1800,
        "CU0": 78000,
        "AU0": 560,
        "M0": 3100,
        "CF0": 15000,
    }.get(symbol, 3000)
    drift = rng.normal(0.0002, 0.0005)
    vol = rng.uniform(0.008, 0.025)
    ret = rng.normal(drift, vol, size=n)
    close = base * np.cumprod(1 + ret)
    open_ = np.r_[close[0] * (1 + rng.normal(0, vol / 2)), close[:-1] * (1 + rng.normal(0, vol / 3, size=n - 1))]
    high = np.maximum(open_, close) * (1 + np.abs(rng.normal(0, vol / 2, size=n)))
    low = np.minimum(open_, close) * (1 - np.abs(rng.normal(0, vol / 2, size=n)))
    volume = rng.integers(20000, 500000, size=n)
    hold = rng.integers(30000, 600000, size=n)
    settle = (open_ + high + low + close) / 4
    return pd.DataFrame({
        "date": dates,
        "open": open_,
        "high": high,
        "low": low,
        "close": close,
        "volume": volume,
        "hold": hold,
        "settle": settle,
    })


def fetch_daily(symbol: str, mock: bool = False) -> pd.DataFrame:
    if mock:
        return generate_mock_daily(symbol)
    ak = get_akshare()
    errors = []
    # 第一优先：新浪日频，连续合约 symbol = 品种代码 + 0
    try:
        df = ak.futures_zh_daily_sina(symbol=symbol)
        out = normalize_daily_df(df)
        if len(out) >= 30:
            return out
    except Exception as e:
        errors.append(f"futures_zh_daily_sina失败: {e}")
    # 第二优先：主力连续历史接口（新浪）
    try:
        end = datetime.today().strftime("%Y%m%d")
        start = (pd.Timestamp.today() - pd.Timedelta(days=800)).strftime("%Y%m%d")
        df = ak.futures_main_sina(symbol=symbol, start_date=start, end_date=end)
        out = normalize_daily_df(df)
        if len(out) >= 30:
            return out
    except Exception as e:
        errors.append(f"futures_main_sina失败: {e}")
    # 第三优先：东方财富接口（海外可用）
    try:
        base = symbol_base(symbol).lower()
        end = datetime.today().strftime("%Y%m%d")
        start = (pd.Timestamp.today() - pd.Timedelta(days=800)).strftime("%Y%m%d")
        df = ak.futures_zh_hist(symbol=base, period="daily", start_date=start, end_date=end, adjust="")
        out = normalize_daily_df(df)
        if len(out) >= 30:
            return out
    except Exception as e:
        errors.append(f"futures_zh_hist(东财)失败: {e}")
    raise RuntimeError("; ".join(errors))


def update_one_symbol(symbol: str, name: str, exchange: str, config: dict, mock: bool = False) -> Tuple[pd.DataFrame, dict]:
    db_dir = Path(config["database_dir"])
    csv_path = db_dir / f"{symbol}.csv"
    status = {
        "symbol": symbol,
        "name": clean_name(name),
        "exchange": exchange,
        "ok": False,
        "source": "network" if not mock else "mock",
        "message": "",
        "last_date": "",
        "rows": 0,
        "运行时间": "",
        "数据新鲜度": "",
        "运行日": "",
        "距运行日天数": "",
    }
    try:
        new_df = fetch_daily(symbol, mock=mock)
        if csv_path.exists() and not mock:
            old = normalize_daily_df(pd.read_csv(csv_path))
            df = pd.concat([old, new_df], ignore_index=True)
            df = df.sort_values("date").drop_duplicates("date", keep="last")
        else:
            df = new_df
        df = normalize_daily_df(df)
        keep = int(config.get("lookback_bars", 250)) + 80  # 多保留少量用于因子暖机
        df = df.tail(keep).reset_index(drop=True)
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        status.update({
            "ok": True,
            "message": "更新成功",
            "last_date": df["date"].max().strftime("%Y-%m-%d") if not df.empty else "",
            "rows": len(df),
        })
        return df, status
    except Exception as e:
        if csv_path.exists():
            try:
                df = normalize_daily_df(pd.read_csv(csv_path))
                status.update({
                    "ok": True,
                    "source": "local_cache",
                    "message": f"联网失败，使用本地缓存：{e}",
                    "last_date": df["date"].max().strftime("%Y-%m-%d") if not df.empty else "",
                    "rows": len(df),
                })
                return df, status
            except Exception:
                pass
        status["message"] = str(e)
        return pd.DataFrame(), status


# -----------------------------
# 指标计算
# -----------------------------

def rolling_atr(df: pd.DataFrame, window: int = 14) -> pd.Series:
    prev_close = df["close"].shift(1)
    tr = pd.concat([
        df["high"] - df["low"],
        (df["high"] - prev_close).abs(),
        (df["low"] - prev_close).abs(),
    ], axis=1).max(axis=1)
    return tr.rolling(window, min_periods=max(3, window // 2)).mean()


def detect_confirmed_pivots(df: pd.DataFrame) -> Tuple[pd.Series, pd.Series]:
    # 当天高点高于前5日和后5日，需后5日确认。为避免未来函数，当前可用信号 shift(5)。
    high = df["high"]
    low = df["low"]
    left_high = high.rolling(6, min_periods=6).max()
    right_high = high[::-1].rolling(6, min_periods=6).max()[::-1]
    left_low = low.rolling(6, min_periods=6).min()
    right_low = low[::-1].rolling(6, min_periods=6).min()[::-1]
    pivot_high_raw = (high >= left_high) & (high >= right_high)
    pivot_low_raw = (low <= left_low) & (low <= right_low)
    # 到 t 日只能确认 t-5 的 pivot
    pivot_high = pivot_high_raw.shift(5).fillna(False)
    pivot_low = pivot_low_raw.shift(5).fillna(False)
    return pivot_high.astype(bool), pivot_low.astype(bool)


def round_step(price: float) -> float:
    p = abs(price)
    if p < 100:
        return 5
    if p < 500:
        return 10
    if p < 1000:
        return 50
    if p < 3000:
        return 100
    if p < 10000:
        return 500
    if p < 30000:
        return 1000
    return 5000


def lin_value(x1: int, y1: float, x2: int, y2: float, x: int) -> float:
    if x2 == x1:
        return y2
    return y1 + (y2 - y1) * (x - x1) / (x2 - x1)


def calculate_factors(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy().reset_index(drop=True)
    n = len(df)
    if n < 80:
        return pd.DataFrame()

    df["ret_cc"] = df["close"].pct_change()
    df["next_intraday_ret"] = df["close"].shift(-1) / df["open"].shift(-1) - 1
    df["range"] = df["high"] - df["low"]
    df["body"] = df["close"] - df["open"]
    df["atr7"] = df["range"].rolling(7, min_periods=5).mean()
    df["atr14"] = rolling_atr(df, 14)
    df["ma5"] = df["close"].rolling(5, min_periods=3).mean()
    df["ma20"] = df["close"].rolling(20, min_periods=10).mean()
    df["ma60"] = df["close"].rolling(60, min_periods=30).mean()
    df["vol20"] = df["volume"].rolling(20, min_periods=10).mean()
    df["hold_chg"] = df["hold"].diff()
    df["hold_chg20"] = df["hold_chg"].rolling(20, min_periods=10).mean()
    df["high20"] = df["high"].rolling(20, min_periods=10).max()
    df["low20"] = df["low"].rolling(20, min_periods=10).min()
    df["high60"] = df["high"].rolling(60, min_periods=30).max()
    df["low60"] = df["low"].rolling(60, min_periods=30).min()
    df["atr_pct"] = df["atr14"] / df["close"]
    df["atr_pct_ma20"] = df["atr_pct"].rolling(20, min_periods=10).mean()

    pivot_high, pivot_low = detect_confirmed_pivots(df)
    df["pivot_high"] = pivot_high
    df["pivot_low"] = pivot_low

    factors = pd.DataFrame(index=df.index)
    for f in FACTOR_NAMES:
        factors[f] = 0

    # 1. 前高前低：突破/跌破已确认 pivot，或靠近支撑压力
    last_ph = np.nan
    last_pl = np.nan
    ph_list: List[Tuple[int, float]] = []
    pl_list: List[Tuple[int, float]] = []
    f1 = []
    f4 = []
    for i, row in df.iterrows():
        if bool(row["pivot_high"]):
            idx = max(0, i - 5)
            ph_list.append((idx, float(df.loc[idx, "high"])))
            last_ph = ph_list[-1][1]
        if bool(row["pivot_low"]):
            idx = max(0, i - 5)
            pl_list.append((idx, float(df.loc[idx, "low"])))
            last_pl = pl_list[-1][1]

        close = row["close"]
        atr = row["atr14"] if not pd.isna(row["atr14"]) else row["range"]
        buf = max(0.003 * close, 0.3 * atr if not pd.isna(atr) else 0)
        sig1 = 0
        if not pd.isna(last_ph) and close > last_ph + buf:
            sig1 = 1
        elif not pd.isna(last_pl) and close < last_pl - buf:
            sig1 = -1
        elif not pd.isna(last_pl) and abs(close - last_pl) <= buf:
            sig1 = 1
        elif not pd.isna(last_ph) and abs(close - last_ph) <= buf:
            sig1 = -1
        f1.append(sig1)

        # 4. 颈线趋势线：最后两个 pivot 高/低连接
        sig4 = 0
        if len(ph_list) >= 2:
            (x1, y1), (x2, y2) = ph_list[-2], ph_list[-1]
            line = lin_value(x1, y1, x2, y2, i)
            if close > line + buf and y2 < y1:  # 突破下降压力线
                sig4 = 1
        if len(pl_list) >= 2:
            (x1, y1), (x2, y2) = pl_list[-2], pl_list[-1]
            line = lin_value(x1, y1, x2, y2, i)
            if close < line - buf and y2 > y1:  # 跌破上升支撑线
                sig4 = -1
        f4.append(sig4)

    factors["前高前低"] = f1
    factors["颈线趋势线"] = f4

    # 2. 大阳大阴 + 半分位
    big_bull = df["body"] > 2 * df["atr7"]
    big_bear = -df["body"] > 2 * df["atr7"]
    f2 = np.zeros(n, dtype=int)
    f2[big_bull.fillna(False).values] = 1
    f2[big_bear.fillna(False).values] = -1
    last_big_yang_mid = np.nan
    last_big_yin_mid = np.nan
    for i in range(n):
        if bool(big_bull.iloc[i]):
            last_big_yang_mid = (df.loc[i, "open"] + df.loc[i, "close"]) / 2
        if bool(big_bear.iloc[i]):
            last_big_yin_mid = (df.loc[i, "open"] + df.loc[i, "close"]) / 2
        if f2[i] == 0:
            close = df.loc[i, "close"]
            if not pd.isna(last_big_yang_mid) and close >= last_big_yang_mid and abs(close - last_big_yang_mid) / close < 0.01:
                f2[i] = 1
            elif not pd.isna(last_big_yin_mid) and close <= last_big_yin_mid and abs(close - last_big_yin_mid) / close < 0.01:
                f2[i] = -1
    factors["大阳大阴"] = f2

    # 3. 跳空缺口
    up_gap = df["low"] > df["high"].shift(1)
    down_gap = df["high"] < df["low"].shift(1)
    factors.loc[up_gap.fillna(False), "跳空缺口"] = 1
    factors.loc[down_gap.fillna(False), "跳空缺口"] = -1

    # 5. 整数关口
    f6 = []
    prev_close = df["close"].shift(1)
    for i, row in df.iterrows():
        close = row["close"]
        step = round_step(close)
        level = round(close / step) * step
        buf = max(0.002 * close, 0.2 * (row["atr14"] if not pd.isna(row["atr14"]) else row["range"]))
        sig = 0
        pc = prev_close.iloc[i]
        if not pd.isna(pc):
            if pc < level <= close and abs(close - level) <= 2 * buf:
                sig = 1
            elif pc > level >= close and abs(close - level) <= 2 * buf:
                sig = -1
            elif abs(close - level) <= buf:
                sig = 1 if close >= level else -1
        f6.append(sig)
    factors["整数关口"] = f6

    # 6. 黄金分割：近60日高低点
    f7 = np.zeros(n, dtype=int)
    prev = df["close"].shift(1)
    for i in range(n):
        hi = df.loc[i, "high60"]
        lo = df.loc[i, "low60"]
        if pd.isna(hi) or pd.isna(lo) or hi <= lo or pd.isna(prev.iloc[i]):
            continue
        levels = [lo + (hi - lo) * r for r in [0.382, 0.5, 0.618]]
        for lv in levels:
            if prev.iloc[i] < lv <= df.loc[i, "close"]:
                f7[i] = 1
            elif prev.iloc[i] > lv >= df.loc[i, "close"]:
                f7[i] = -1
    factors["黄金分割"] = f7

    # 7. 趋势方向
    factors.loc[((df["close"] > df["ma20"]) & (df["ma20"] > df["ma60"])).fillna(False), "趋势方向"] = 1
    factors.loc[((df["close"] < df["ma20"]) & (df["ma20"] < df["ma60"])).fillna(False), "趋势方向"] = -1

    # 8. 支撑阻力
    prev_high20 = df["high20"].shift(1)
    prev_low20 = df["low20"].shift(1)
    factors.loc[(df["close"] > prev_high20).fillna(False), "支撑阻力"] = 1
    factors.loc[(df["close"] < prev_low20).fillna(False), "支撑阻力"] = -1

    # 9. 成交量确认
    vol_boom = df["volume"] > 1.3 * df["vol20"]
    factors.loc[((df["ret_cc"] > 0) & vol_boom).fillna(False), "成交量确认"] = 1
    factors.loc[((df["ret_cc"] < 0) & vol_boom).fillna(False), "成交量确认"] = -1

    # 10. 持仓量确认
    hold_inc = df["hold_chg"] > df["hold_chg20"].abs().fillna(0)
    factors.loc[((df["ret_cc"] > 0) & hold_inc).fillna(False), "持仓量确认"] = 1
    factors.loc[((df["ret_cc"] < 0) & hold_inc).fillna(False), "持仓量确认"] = -1

    # 11. 移动均线
    cross_up = (df["ma5"] > df["ma20"]) & (df["ma5"].shift(1) <= df["ma20"].shift(1))
    cross_dn = (df["ma5"] < df["ma20"]) & (df["ma5"].shift(1) >= df["ma20"].shift(1))
    factors.loc[cross_up.fillna(False), "移动均线"] = 1
    factors.loc[cross_dn.fillna(False), "移动均线"] = -1
    factors.loc[((factors["移动均线"] == 0) & (df["close"] > df["ma5"]) & (df["ma5"] > df["ma20"])).fillna(False), "移动均线"] = 1
    factors.loc[((factors["移动均线"] == 0) & (df["close"] < df["ma5"]) & (df["ma5"] < df["ma20"])).fillna(False), "移动均线"] = -1

    # 12. 形态突破：箱体压缩后突破
    range20_pct = (df["high20"] - df["low20"]) / df["close"]
    compressed = range20_pct < range20_pct.rolling(120, min_periods=50).quantile(0.35)
    factors.loc[((compressed.shift(1)) & (df["close"] > prev_high20)).fillna(False), "形态突破"] = 1
    factors.loc[((compressed.shift(1)) & (df["close"] < prev_low20)).fillna(False), "形态突破"] = -1

    # 13. 波动率变化
    vol_expand = df["atr_pct"] > 1.2 * df["atr_pct_ma20"]
    factors.loc[((vol_expand) & (df["ret_cc"] > 0)).fillna(False), "波动率变化"] = 1
    factors.loc[((vol_expand) & (df["ret_cc"] < 0)).fillna(False), "波动率变化"] = -1

    factors = factors.fillna(0).astype(int)
    df["bull_count"] = (factors == 1).sum(axis=1)
    df["bear_count"] = (factors == -1).sum(axis=1)
    df["net_factor_score"] = df["bull_count"] - df["bear_count"]
    out = pd.concat([df, factors.add_prefix("因子_")], axis=1)
    return out


# -----------------------------
# 回测与排序
# -----------------------------

def signal_from_threshold(row: pd.Series, min_factors: int, net_gap: int) -> int:
    bull = int(row.get("bull_count", 0))
    bear = int(row.get("bear_count", 0))
    net = bull - bear
    if bull >= min_factors and net >= net_gap:
        return 1
    if bear >= min_factors and -net >= net_gap:
        return -1
    return 0


def evaluate_threshold(df: pd.DataFrame, min_factors: int, net_gap: int, weights: dict) -> Optional[dict]:
    tmp = df.copy()
    tmp["signal"] = tmp.apply(lambda r: signal_from_threshold(r, min_factors, net_gap), axis=1)
    tmp = tmp.dropna(subset=["next_intraday_ret"])
    trades = tmp[tmp["signal"] != 0].copy()
    if trades.empty:
        return None
    trades["strategy_ret"] = trades["signal"] * trades["next_intraday_ret"]
    win_rate = float((trades["strategy_ret"] > 0).mean())
    avg_ret = float(trades["strategy_ret"].mean())
    pf = profit_factor(trades["strategy_ret"])
    max_dd = calc_max_drawdown(trades["strategy_ret"])
    long_trades = int((trades["signal"] == 1).sum())
    short_trades = int((trades["signal"] == -1).sum())
    bias = "偏多" if long_trades > short_trades else "偏空" if short_trades > long_trades else "均衡"
    comp = score_model(win_rate, avg_ret, pf, max_dd, len(trades), weights)
    return {
        "min_factors": min_factors,
        "net_gap": net_gap,
        "trades": int(len(trades)),
        "win_rate": win_rate,
        "avg_return": avg_ret,
        "profit_factor": pf,
        "max_drawdown": max_dd,
        "composite_score": comp,
        "direction_bias": bias,
        "long_trades": long_trades,
        "short_trades": short_trades,
    }


def find_best_models(symbol: str, name: str, df: pd.DataFrame, config: dict) -> Tuple[Optional[BestModel], Optional[BestModel], pd.DataFrame]:
    weights = config.get("score_weights", {})
    ranges = config.get("factor_thresholds", {})
    min_factors_range = ranges.get("min_factors_range", [3, 4, 5, 6, 7, 8, 9, 10])
    net_gap_range = ranges.get("net_gap_range", [1, 2, 3, 4])
    results = []
    # 回测只使用最近 lookback_bars 天，但要排除最后一天，因为没有下一日结果
    lookback = int(config.get("lookback_bars", 250))
    bt_df = df.tail(lookback + 1).copy()
    for mf in min_factors_range:
        for gap in net_gap_range:
            res = evaluate_threshold(bt_df, int(mf), int(gap), weights)
            if res is not None:
                res["symbol"] = symbol
                res["name"] = name
                results.append(res)
    if not results:
        return None, None, pd.DataFrame()
    res_df = pd.DataFrame(results)
    regular_min = int(config.get("regular_min_trades", 20))
    lowfreq_min = int(config.get("lowfreq_min_trades", 5))
    regular = res_df[res_df["trades"] >= regular_min].sort_values("composite_score", ascending=False)
    lowfreq = res_df[(res_df["trades"] >= lowfreq_min) & (res_df["trades"] < regular_min)].sort_values(
        ["win_rate", "composite_score"], ascending=[False, False]
    )

    def to_model(row: pd.Series, model_type: str) -> BestModel:
        return BestModel(
            symbol=symbol,
            name=name,
            min_factors=int(row["min_factors"]),
            net_gap=int(row["net_gap"]),
            trades=int(row["trades"]),
            win_rate=float(row["win_rate"]),
            avg_return=float(row["avg_return"]),
            profit_factor=float(row["profit_factor"]),
            max_drawdown=float(row["max_drawdown"]),
            composite_score=float(row["composite_score"]),
            direction_bias=str(row["direction_bias"]),
            model_type=model_type,
        )

    best_regular = to_model(regular.iloc[0], "常规模型") if not regular.empty else None
    best_lowfreq = to_model(lowfreq.iloc[0], "高胜率低频") if not lowfreq.empty else None
    return best_regular, best_lowfreq, res_df


def latest_signal(df: pd.DataFrame, model: Optional[BestModel]) -> Tuple[str, int, int, int]:
    if df.empty:
        return "无数据", 0, 0, 0
    row = df.iloc[-1]
    bull = int(row.get("bull_count", 0))
    bear = int(row.get("bear_count", 0))
    net = bull - bear
    if model is None:
        if net >= 3:
            return "观察偏多", 0, bull, bear
        if net <= -3:
            return "观察偏空", 0, bull, bear
        return "观望", 0, bull, bear
    sig = signal_from_threshold(row, model.min_factors, model.net_gap)
    if sig > 0:
        return "建议做多", sig, bull, bear
    if sig < 0:
        return "建议做空", sig, bull, bear
    if net >= 3:
        return "观察偏多", 0, bull, bear
    if net <= -3:
        return "观察偏空", 0, bull, bear
    return "观望", 0, bull, bear


def current_bias_text(advice: str, sig: int, bull: int, bear: int) -> str:
    """当前最新一日信号方向。

    v1.1 中“方向偏好”使用的是历史回测中多空信号次数的偏好，
    会与“今日建议”出现看似相反的情况。v1.3 改为：
    - 方向偏好：只表示今天这根K线触发后的当前方向；
    - 历史回测偏好：另列显示历史模型更常触发多头还是空头。
    """
    if sig > 0 or "做多" in str(advice) or "偏多" in str(advice):
        return "偏多"
    if sig < 0 or "做空" in str(advice) or "偏空" in str(advice):
        return "偏空"
    if bull > bear:
        return "偏多观察"
    if bear > bull:
        return "偏空观察"
    return "中性"


# -----------------------------
# 合约规格/保证金：尽力获取，失败则留空
# -----------------------------

def parse_multiplier(text: str) -> Optional[float]:
    if not text:
        return None
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)", str(text))
    if m:
        return float(m.group(1))
    return None


def parse_margin_ratio(text: str) -> Optional[float]:
    if not text:
        return None
    vals = re.findall(r"([0-9]+(?:\.[0-9]+)?)\s*%", str(text))
    if vals:
        return float(vals[0]) / 100
    return None


def fetch_contract_specs(symbols: List[str], mock: bool = False) -> Dict[str, dict]:
    specs: Dict[str, dict] = {}
    if mock:
        for s in symbols:
            specs[s] = {"multiplier": 10, "margin_ratio": 0.12, "fee": "模拟", "main_contract": s, "spec_message": "模拟规格"}
        return specs
    try:
        ak = get_akshare()
    except Exception:
        return specs
    for s in symbols:
        base = symbol_base(s)
        specs[s] = {"multiplier": np.nan, "margin_ratio": np.nan, "fee": "", "main_contract": s, "spec_message": ""}
        candidates = [s]
        # 连续合约详情有时不返回，尝试用当前主力合约字符串补充。
        try:
            # 不同版本接口返回可能是字符串或列表，这里只做最小依赖。
            pass
        except Exception:
            pass
        ok = False
        for c in candidates:
            try:
                det = ak.futures_contract_detail(symbol=c)
                if det is None or det.empty:
                    continue
                item_col = "item" if "item" in det.columns else det.columns[0]
                val_col = "value" if "value" in det.columns else det.columns[1]
                mp = dict(zip(det[item_col].astype(str), det[val_col].astype(str)))
                unit = mp.get("交易单位", "")
                margin = mp.get("最低交易保证金", "")
                fee = mp.get("交易手续费", "")
                specs[s].update({
                    "multiplier": parse_multiplier(unit),
                    "margin_ratio": parse_margin_ratio(margin),
                    "fee": fee,
                    "main_contract": c,
                    "spec_message": "已获取",
                })
                ok = True
                break
            except Exception as e:
                specs[s]["spec_message"] = str(e)[:80]
        if not ok:
            specs[s]["spec_message"] = specs[s].get("spec_message") or "未获取到合约规格"
    return specs



# -----------------------------
# 盘中实时快照与振幅统计
# -----------------------------

def is_intraday_realtime_window(run_ts: datetime, config: dict) -> bool:
    """判断是否处于需要使用实时快照的交易时段。

    这里只做通用窗口判断：日盘 + 夜盘集合竞价/交易时段。
    各品种夜盘结束时间不同，若实时接口返回不到有效数据，会自动退回日线模式。
    """
    if not bool(config.get("enable_intraday_realtime_snapshot", True)):
        return False
    # 周末一般无交易，但周五夜盘/节假日特殊安排不在这里硬编码。
    if run_ts.weekday() >= 5:
        return False
    t = run_ts.time()
    windows = [
        (time(8, 55), time(10, 15)),
        (time(10, 30), time(11, 30)),
        (time(13, 30), time(15, 0)),
        (time(20, 55), time(23, 59, 59)),
        (time(0, 0), time(2, 30)),
    ]
    return any(a <= t <= b for a, b in windows)


def calc_range_stats_from_daily(df: pd.DataFrame, window: int = 10) -> dict:
    """近N个完整日线的价格振幅统计。振幅=high-low。"""
    if df is None or df.empty:
        return {"avg_range10": np.nan, "min_range10": np.nan, "max_range10": np.nan, "avg_range10_pct": np.nan, "min_range10_pct": np.nan}
    tmp = normalize_daily_df(df).tail(window).copy()
    if tmp.empty:
        return {"avg_range10": np.nan, "min_range10": np.nan, "max_range10": np.nan, "avg_range10_pct": np.nan, "min_range10_pct": np.nan}
    tmp["range"] = tmp["high"] - tmp["low"]
    denom = tmp["settle"].replace(0, np.nan)
    tmp["range_pct"] = tmp["range"] / denom
    return {
        "avg_range10": float(tmp["range"].mean()),
        "min_range10": float(tmp["range"].min()),
        "max_range10": float(tmp["range"].max()),
        "avg_range10_pct": float(tmp["range_pct"].mean()),
        "min_range10_pct": float(tmp["range_pct"].min()),
    }


def extract_contract_code(text: str) -> Optional[str]:
    """从主力合约订阅文本或行情名称中提取合约代码。"""
    if not text:
        return None
    raw = str(text).upper().replace("NF_", "").replace("HF_", "")
    m = re.search(r"([A-Z]{1,3}[0-9]{3,4})", raw)
    if m:
        return m.group(1)
    return None


def build_main_contract_map(mock: bool = False) -> Dict[str, str]:
    """返回 品种字母代码 -> 当前主力合约代码，例如 L -> L2609。"""
    if mock:
        return {"RB": "RB2609", "LH": "LH2609", "I": "I2609", "SA": "SA2609", "CU": "CU2609", "AU": "AU2608", "M": "M2609", "CF": "CF2609"}
    try:
        ak = get_akshare()
    except Exception:
        return {}
    out: Dict[str, str] = {}
    for ex in ["dce", "czce", "shfe", "gfex"]:
        try:
            text = ak.match_main_contract(symbol=ex)
            # 返回通常是逗号分隔的合约代码字符串，可直接订阅 futures_zh_spot。
            for part in re.split(r"[,，\s]+", str(text)):
                code = extract_contract_code(part)
                if not code:
                    continue
                base = re.match(r"([A-Z]{1,3})", code).group(1)
                out[base] = code
        except Exception:
            continue
    return out


def normalize_spot_df(df: pd.DataFrame, requested_codes: List[str]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    # 部分版本列名为中文，这里统一成英文。
    rename = {
        "品种": "symbol",
        "时间": "time",
        "开盘": "open",
        "最高": "high",
        "最低": "low",
        "最新价": "current_price",
        "当前价格": "current_price",
        "买价": "bid_price",
        "卖价": "ask_price",
        "买": "bid_price",
        "卖": "ask_price",
        "买量": "buy_vol",
        "卖量": "sell_vol",
        "持仓量": "hold",
        "成交量": "volume",
        "均价": "avg_price",
        "昨收": "last_close",
        "上一个交易日的收盘价": "last_close",
        "昨结算": "last_settle_price",
        "上一个交易日的结算价": "last_settle_price",
    }
    out.rename(columns=rename, inplace=True)
    if "requested_contract" not in out.columns:
        # futures_zh_spot 返回顺序通常与请求顺序一致。若行数不一致，后面仍会用 symbol 尝试解析。
        vals = requested_codes[: len(out)] + [""] * max(0, len(out) - len(requested_codes))
        out["requested_contract"] = vals[: len(out)]
    need_cols = ["symbol", "time", "open", "high", "low", "current_price", "bid_price", "ask_price", "hold", "volume", "avg_price", "last_close", "last_settle_price", "requested_contract"]
    for c in need_cols:
        if c not in out.columns:
            out[c] = np.nan
    for c in ["open", "high", "low", "current_price", "bid_price", "ask_price", "hold", "volume", "avg_price", "last_close", "last_settle_price"]:
        out[c] = pd.to_numeric(out[c], errors="coerce")
    return out[need_cols]


def fetch_realtime_snapshots(active_symbols: List[str], data_map: Dict[str, pd.DataFrame], config: dict, mock: bool = False) -> Tuple[Dict[str, dict], pd.DataFrame, str]:
    """获取主力合约实时快照。

    返回：
    - live_map: 连续合约代码 -> 实时字段
    - live_df: 用于输出的实时快照表
    - message: 状态说明
    """
    if not bool(config.get("enable_intraday_realtime_snapshot", True)):
        return {}, pd.DataFrame(), "配置关闭实时快照"
    if mock:
        rows = []
        live_map = {}
        rng = np.random.default_rng(202406)
        for symbol in active_symbols:
            df = data_map.get(symbol, pd.DataFrame())
            if df.empty:
                continue
            last = df.iloc[-1]
            base = symbol_base(symbol)
            contract = f"{base}2609"
            cur = float(last["close"] * (1 + rng.normal(0, 0.003)))
            op = float(last["close"] * (1 + rng.normal(0, 0.002)))
            hi = max(cur, op) * (1 + abs(rng.normal(0, 0.004)))
            lo = min(cur, op) * (1 - abs(rng.normal(0, 0.004)))
            rec = {
                "symbol": symbol,
                "contract": contract,
                "spot_symbol": contract,
                "spot_time": datetime.now().strftime("%H%M%S"),
                "open": op,
                "high": hi,
                "low": lo,
                "current_price": cur,
                "bid_price": cur - 1,
                "ask_price": cur + 1,
                "hold": float(last.get("hold", 0)),
                "volume": float(last.get("volume", 0)),
                "avg_price": (op + hi + lo + cur) / 4,
                "last_close": float(last.get("close", np.nan)),
                "last_settle_price": float(last.get("settle", np.nan)),
                "source": "mock_realtime",
            }
            live_map[symbol] = rec
            rows.append(rec)
        return live_map, pd.DataFrame(rows), "模拟实时快照"

    try:
        ak = get_akshare()
    except Exception as e:
        return {}, pd.DataFrame(), f"无法导入akshare，跳过实时快照：{e}"

    main_map = build_main_contract_map(mock=False)
    requested: List[str] = []
    symbol_for_contract: Dict[str, str] = {}
    for symbol in active_symbols:
        base = symbol_base(symbol)
        contract = main_map.get(base)
        if contract:
            requested.append(contract)
            symbol_for_contract[contract.upper()] = symbol
    if not requested:
        return {}, pd.DataFrame(), "未获取到主力合约映射，跳过实时快照"

    try:
        spot = ak.futures_zh_spot(symbol=",".join(requested), market="CF", adjust="0")
        spot = normalize_spot_df(spot, requested)
    except Exception as e:
        return {}, pd.DataFrame(), f"实时行情接口失败：{e}"

    rows = []
    live_map: Dict[str, dict] = {}
    for idx, row in spot.reset_index(drop=True).iterrows():
        req = str(row.get("requested_contract", "")).upper()
        parsed = extract_contract_code(str(row.get("symbol", ""))) or req
        parsed = str(parsed).upper()
        contract = parsed if parsed else req
        symbol = symbol_for_contract.get(contract) or symbol_for_contract.get(req)
        if not symbol:
            base = re.match(r"([A-Z]{1,3})", contract)
            if base:
                symbol = f"{base.group(1)}0"
        if not symbol or symbol not in active_symbols:
            continue
        cur = safe_float(row.get("current_price"))
        bid = safe_float(row.get("bid_price"))
        ask = safe_float(row.get("ask_price"))
        # current_price有时取买价，若为空则用买卖中间价。
        if pd.isna(cur) and not pd.isna(bid) and not pd.isna(ask):
            cur = (bid + ask) / 2
        if pd.isna(cur) or cur <= 0:
            continue
        rec = {
            "symbol": symbol,
            "contract": contract,
            "spot_symbol": row.get("symbol", ""),
            "spot_time": str(row.get("time", "")),
            "open": safe_float(row.get("open")),
            "high": safe_float(row.get("high")),
            "low": safe_float(row.get("low")),
            "current_price": cur,
            "bid_price": bid,
            "ask_price": ask,
            "hold": safe_float(row.get("hold")),
            "volume": safe_float(row.get("volume")),
            "avg_price": safe_float(row.get("avg_price")),
            "last_close": safe_float(row.get("last_close")),
            "last_settle_price": safe_float(row.get("last_settle_price")),
            "source": "futures_zh_spot",
        }
        live_map[symbol] = rec
        rows.append(rec)
    msg = f"实时快照成功：{len(live_map)}/{len(active_symbols)}个活跃品种"
    return live_map, pd.DataFrame(rows), msg


def append_intraday_snapshot(raw: pd.DataFrame, live: Optional[dict], run_ts: datetime, config: dict) -> Tuple[pd.DataFrame, str]:
    """盘中把实时快照临时追加为最新一行，仅用于当前信号和看板，不写入数据库。"""
    if raw is None or raw.empty:
        return raw, "无日线数据"
    if not is_intraday_realtime_window(run_ts, config) or not live:
        return raw, "盘后日线"
    cur = safe_float(live.get("current_price"))
    op = safe_float(live.get("open"))
    hi = safe_float(live.get("high"))
    lo = safe_float(live.get("low"))
    if pd.isna(cur) or cur <= 0:
        return raw, "实时快照无有效价格，使用日线"
    last = raw.iloc[-1]
    if pd.isna(op) or op <= 0:
        op = float(last.get("close", cur))
    if pd.isna(hi) or hi <= 0:
        hi = max(op, cur)
    if pd.isna(lo) or lo <= 0:
        lo = min(op, cur)
    if hi < max(op, cur):
        hi = max(op, cur)
    if lo > min(op, cur):
        lo = min(op, cur)

    latest_daily_date = pd.Timestamp(last["date"]).normalize()
    if run_ts.time() >= time(20, 55) and latest_daily_date.date() == run_ts.date():
        synth_date = latest_daily_date + pd.offsets.BDay(1)
    else:
        synth_date = pd.Timestamp(run_ts.date())
        if synth_date <= latest_daily_date:
            synth_date = latest_daily_date + pd.offsets.BDay(1)

    live_row = {
        "date": synth_date,
        "open": op,
        "high": hi,
        "low": lo,
        "close": cur,
        "volume": safe_float(live.get("volume"), 0),
        "hold": safe_float(live.get("hold"), last.get("hold", 0)),
        "settle": safe_float(live.get("last_settle_price"), last.get("settle", cur)),
    }
    out = pd.concat([raw, pd.DataFrame([live_row])], ignore_index=True)
    out = normalize_daily_df(out)
    return out, "盘中实时快照"


def live_fields(live: Optional[dict], mode: str) -> dict:
    if not live:
        return {
            "行情模式": mode,
            "实时合约": "",
            "实时行情时间": "",
            "实时价格": np.nan,
            "实时开盘": np.nan,
            "实时最高": np.nan,
            "实时最低": np.nan,
            "盘中振幅": np.nan,
            "盘中振幅率": np.nan,
            "实时成交量": np.nan,
            "实时持仓量": np.nan,
        }
    hi = safe_float(live.get("high"))
    lo = safe_float(live.get("low"))
    last_settle = safe_float(live.get("last_settle_price"))
    intraday_range = hi - lo if not any(pd.isna(x) for x in [hi, lo]) else np.nan
    intraday_range_pct = intraday_range / last_settle if not any(pd.isna(x) for x in [intraday_range, last_settle]) and last_settle != 0 else np.nan
    return {
        "行情模式": mode,
        "实时合约": live.get("contract", ""),
        "实时行情时间": live.get("spot_time", ""),
        "实时价格": safe_float(live.get("current_price")),
        "实时开盘": safe_float(live.get("open")),
        "实时最高": hi,
        "实时最低": lo,
        "盘中振幅": intraday_range,
        "盘中振幅率": intraday_range_pct,
        "实时成交量": safe_float(live.get("volume")),
        "实时持仓量": safe_float(live.get("hold")),
    }

# -----------------------------
# 报表输出
# -----------------------------

def pct(x: float) -> str:
    if pd.isna(x):
        return ""
    return f"{x:.2%}"


def build_report_rows(
    all_results: List[dict],
    lowfreq_rows: List[dict],
    data_status: List[dict],
    factor_latest_rows: List[dict],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    all_df = pd.DataFrame(all_results)
    low_df = pd.DataFrame(lowfreq_rows)
    status_df = pd.DataFrame(data_status)
    factor_df = pd.DataFrame(factor_latest_rows)
    if not all_df.empty:
        all_df = all_df.sort_values(["是否建议", "综合评分", "信号强度"], ascending=[False, False, False])
    if not low_df.empty:
        low_df = low_df.sort_values(["今日信号分", "胜率", "综合评分"], ascending=[False, False, False])
    return all_df, low_df, status_df, factor_df


def export_excel(
    path: Path,
    dashboard_df: pd.DataFrame,
    all_df: pd.DataFrame,
    low_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    factor_df: pd.DataFrame,
    status_df: pd.DataFrame,
    realtime_df: pd.DataFrame,
    config: dict,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.formatting.rule import CellIsRule, DataBarRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "今日看板"

    def write_df(sheet, df: pd.DataFrame, title: str):
        sheet.append([title])
        title_row = sheet.max_row
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max(1, len(df.columns)))
        sheet.cell(title_row, 1).font = Font(size=16, bold=True, color="FFFFFF")
        sheet.cell(title_row, 1).fill = PatternFill("solid", fgColor="1F4E78")
        sheet.cell(title_row, 1).alignment = Alignment(horizontal="center")
        sheet.append(list(df.columns) if not df.empty else ["无数据"])
        header_row = sheet.max_row
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(header_row, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if not df.empty:
            for row in df.itertuples(index=False):
                sheet.append(list(row))
        return title_row, header_row

    sheets = [
        ("今日看板", dashboard_df, "明日交易看板 - 前五名与观察池"),
        ("全部品种评分", all_df, "全部活跃品种评分"),
        ("高胜率低频", low_df, "高胜率低频机会池"),
        ("回测明细", backtest_df, "回测参数与绩效明细"),
        ("因子明细", factor_df, "最新交易日 13 因子明细"),
        ("数据状态", status_df, "数据更新状态"),
        ("实时快照", realtime_df, "盘中实时快照（仅用于当前价格/盘中信号，不写入日线数据库）"),
    ]

    for idx, (name, df, title) in enumerate(sheets):
        sheet = ws if idx == 0 else wb.create_sheet(name)
        if idx == 0:
            sheet.title = name
        write_df(sheet, df, title)
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = sheet.dimensions
        thin = Side(style="thin", color="D9E2F3")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        # 列宽
        for col in range(1, sheet.max_column + 1):
            letter = get_column_letter(col)
            max_len = 8
            for cell in sheet[letter]:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except Exception:
                    pass
            sheet.column_dimensions[letter].width = min(max_len + 2, 28)
        # 条件格式
        headers = [sheet.cell(2, c).value for c in range(1, sheet.max_column + 1)] if sheet.max_row >= 2 else []
        for target in ["综合评分", "胜率", "平均收益", "最大回撤"]:
            if target in headers and sheet.max_row > 3:
                c = headers.index(target) + 1
                rng = f"{get_column_letter(c)}3:{get_column_letter(c)}{sheet.max_row}"
                if target == "最大回撤":
                    sheet.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["-0.1"], fill=PatternFill("solid", fgColor="F4CCCC")))
                else:
                    sheet.conditional_formatting.add(rng, DataBarRule(start_type="num", start_value=0, end_type="max", color="63C384"))
        if "建议" in str(headers) or "今日建议" in str(headers):
            for row in range(3, sheet.max_row + 1):
                row_text = " ".join(str(sheet.cell(row, c).value) for c in range(1, sheet.max_column + 1))
                if "做多" in row_text:
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(row, c).fill = PatternFill("solid", fgColor="E2F0D9")
                elif "做空" in row_text:
                    for c in range(1, sheet.max_column + 1):
                        sheet.cell(row, c).fill = PatternFill("solid", fgColor="FCE4D6")

    # 参数说明
    ps = wb.create_sheet("参数说明")
    ps.append(["参数", "当前设置", "说明"])
    for row in PARAM_EXPLAIN:
        ps.append(row)
    ps.append([])
    ps.append(["指标", "说明"])
    for row in FACTOR_EXPLAIN:
        ps.append(row)
    ps.append([])
    ps.append(["配置项", "值"])
    for k, v in config.items():
        ps.append([str(k), json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else str(v)])
    for c in range(1, ps.max_column + 1):
        ps.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ps.cell(1, c).fill = PatternFill("solid", fgColor="1F4E78")
    for col in range(1, ps.max_column + 1):
        ps.column_dimensions[get_column_letter(col)].width = 28 if col == 1 else 70
    for row in ps.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


# -----------------------------
# 静态HTML导出（数据内嵌，无需服务器）
# -----------------------------

def _df_to_records(df: pd.DataFrame) -> list:
    """DataFrame -> JSON友好的records列表。"""
    records = []
    for _, row in df.iterrows():
        record = {}
        for col in df.columns:
            val = row[col]
            if val is None:
                record[col] = None
            elif isinstance(val, float) and (math.isnan(val) if not pd.isna(val) else True):
                record[col] = None
            elif isinstance(val, float):
                record[col] = round(val, 4)
            elif isinstance(val, pd.Timestamp):
                record[col] = val.strftime("%Y-%m-%d %H:%M:%S") if val.hour > 0 else val.strftime("%Y-%m-%d")
            else:
                record[col] = str(val)
        records.append(record)
    return records


_STATIC_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0,maximum-scale=1.0,user-scalable=no">
<title>期货模型交易看板 v1.4 - __FILE_TIME__</title>
<style>
:root{--bg:#0d1117;--card:#161b22;--border:#30363d;--text:#e6edf3;--text2:#8b949e;--blue:#58a6ff;--green:#3fb950;--red:#f85149;--orange:#d29922;--hover:#1c2333}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','PingFang SC','Microsoft YaHei',sans-serif;background:var(--bg);color:var(--text);font-size:14px;line-height:1.5;min-height:100vh;overflow-x:hidden}
.header{background:linear-gradient(135deg,#1a1e2e,#0d1117);border-bottom:1px solid var(--border);padding:12px 16px;position:sticky;top:0;z-index:100;backdrop-filter:blur(10px)}
.header-top{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.header h1{font-size:18px;font-weight:700;color:var(--blue)}
.header h1 small{font-size:12px;color:var(--text2);font-weight:400;margin-left:6px}
.status-bar{display:flex;align-items:center;gap:12px;padding:6px 0 0;font-size:12px;color:var(--text2);overflow-x:auto;white-space:nowrap}
.status-dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:4px}
.tabs{display:flex;gap:2px;overflow-x:auto;padding:8px 16px 0;border-bottom:1px solid var(--border);-webkit-overflow-scrolling:touch}
.tab{padding:8px 14px;border-radius:6px 6px 0 0;cursor:pointer;color:var(--text2);font-size:13px;white-space:nowrap;transition:all .2s;border:1px solid transparent;background:transparent;user-select:none}
.tab:hover{color:var(--text);background:var(--hover)}
.tab.active{color:var(--blue);background:var(--card);border-color:var(--border);border-bottom-color:var(--card);position:relative;top:1px}
.content{padding:12px 16px}
.tab-panel{display:none}.tab-panel.active{display:block;animation:fadeIn .3s ease}
@keyframes fadeIn{from{opacity:0;transform:translateY(10px)}to{opacity:1;transform:translateY(0)}}
.table-wrap{overflow-x:auto;border-radius:8px;border:1px solid var(--border);margin-bottom:16px;background:var(--card)}
.table-wrap .table-title{padding:10px 14px;font-weight:600;font-size:14px;border-bottom:1px solid var(--border);color:var(--blue);background:rgba(88,166,255,.05)}
table{width:100%;border-collapse:collapse;font-size:12px}
th{padding:8px 10px;text-align:left;font-weight:600;background:rgba(88,166,255,.08);color:var(--blue);border-bottom:1px solid var(--border);white-space:nowrap;position:sticky;top:0;z-index:10}
td{padding:6px 10px;border-bottom:1px solid rgba(48,54,61,.5);white-space:nowrap;max-width:200px;overflow:hidden;text-overflow:ellipsis}
tr:hover{background:var(--hover)}
tr.row-long{background:rgba(63,185,80,.08)!important}
tr.row-short{background:rgba(248,81,73,.08)!important}
.badge{display:inline-block;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-long{background:rgba(63,185,80,.15);color:#3fb950}
.badge-short{background:rgba(248,81,73,.15);color:#f85149}
.badge-watch{background:rgba(210,153,34,.15);color:#d29922}
.badge-neutral{background:rgba(139,148,158,.15);color:#8b949e}
.summary-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:10px;margin-bottom:16px}
.summary-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:12px;text-align:center}
.summary-card .label{font-size:11px;color:var(--text2);margin-bottom:4px}
.summary-card .value{font-size:22px;font-weight:700}
.dash-cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:12px;margin-bottom:16px}
.dash-card{background:var(--card);border:1px solid var(--border);border-radius:10px;padding:16px;transition:transform .2s,box-shadow .2s}
.dash-card:hover{transform:translateY(-2px);box-shadow:0 4px 12px rgba(0,0,0,.3)}
.dash-card .dc-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:10px}
.dash-card .dc-symbol{font-size:16px;font-weight:700}
.dash-card .dc-name{font-size:13px;color:var(--text2)}
.dash-card .dc-score{font-size:24px;font-weight:800}
.dash-card .dc-metrics{display:grid;grid-template-columns:1fr 1fr;gap:6px}
.dash-card .dc-metric .dm-label{font-size:11px;color:var(--text2)}
.dash-card .dc-metric .dm-value{font-size:13px;font-weight:600}
.dash-card.is-long{border-left:3px solid var(--green)}
.dash-card.is-short{border-left:3px solid var(--red)}
.dash-card.is-watch{border-left:3px solid var(--orange)}
.search-box{padding:7px 12px;border-radius:6px;border:1px solid var(--border);background:var(--bg);color:var(--text);font-size:13px;outline:none;font-family:inherit}
.search-box:focus{border-color:var(--accent)}
.log-box{background:#0d1117;border:1px solid var(--border);border-radius:6px;padding:10px;font-family:Consolas,Monaco,monospace;font-size:11px;max-height:400px;overflow-y:auto;color:var(--text2);white-space:pre-wrap;word-break:break-all;margin-top:8px}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:var(--border);border-radius:3px}
@media(max-width:768px){.header{padding:10px 12px}.header h1{font-size:15px}.content{padding:8px 10px}.tabs{padding:6px 10px 0}.tab{padding:6px 10px;font-size:12px}td,th{padding:4px 6px;font-size:11px}.dash-cards{grid-template-columns:1fr}.summary-grid{grid-template-columns:repeat(2,1fr)}}
</style>
</head>
<body>
<div class="header">
  <div class="header-top">
    <h1>期货模型交易看板<small>v1.4</small></h1>
    <input type="text" class="search-box" placeholder="搜索品种..." oninput="filterTables(this.value)">
  </div>
  <div class="status-bar">
    <span><span class="status-dot" style="background:var(--green)"></span>数据已嵌入 · 手机直接打开即可</span>
    <span>__FILE_NAME__ (__FILE_TIME__)</span>
  </div>
</div>
<div class="tabs">
  <div class="tab active" onclick="switchTab('dashboard',this)">今日看板</div>
  <div class="tab" onclick="switchTab('all',this)">全部品种</div>
  <div class="tab" onclick="switchTab('lowfreq',this)">高胜率低频</div>
  <div class="tab" onclick="switchTab('backtest',this)">回测明细</div>
  <div class="tab" onclick="switchTab('factors',this)">因子明细</div>
  <div class="tab" onclick="switchTab('status',this)">数据状态</div>
  <div class="tab" onclick="switchTab('realtime',this)">实时快照</div>
  <div class="tab" onclick="switchTab('params',this)">参数说明</div>
</div>
<div class="content">
  <div class="tab-panel active" id="panel-dashboard"></div>
  <div class="tab-panel" id="panel-all"></div>
  <div class="tab-panel" id="panel-lowfreq"></div>
  <div class="tab-panel" id="panel-backtest"></div>
  <div class="tab-panel" id="panel-factors"></div>
  <div class="tab-panel" id="panel-status"></div>
  <div class="tab-panel" id="panel-realtime"></div>
  <div class="tab-panel" id="panel-params"></div>
</div>
<script>
var DATA=__JSON_DATA__;
function switchTab(t,el){document.querySelectorAll('.tab').forEach(function(e){e.classList.remove('active')});el.classList.add('active');document.querySelectorAll('.tab-panel').forEach(function(e){e.classList.toggle('active',e.id==='panel-'+t)})}
function filterTables(kw){kw=kw.toLowerCase();document.querySelectorAll('.table-wrap table tbody tr').forEach(function(tr){tr.style.display=tr.textContent.toLowerCase().includes(kw)?'':'none'})}
function pct(v){return v!=null?(v*100).toFixed(1)+'%':'-'}
function fmt(v,d){d=d||2;return v!=null?Number(v).toFixed(d):'-'}
function badge(t){if(!t)return'<span class="badge badge-neutral">-</span>';if(t.indexOf('做多')>=0)return'<span class="badge badge-long">'+t+'</span>';if(t.indexOf('做空')>=0)return'<span class="badge badge-short">'+t+'</span>';if(t.indexOf('观望')>=0||t.indexOf('观察')>=0)return'<span class="badge badge-watch">'+t+'</span>';return'<span class="badge badge-neutral">'+t+'</span>'}
function rowClass(t){if(!t)return '';if(t.indexOf('做多')>=0)return 'row-long';if(t.indexOf('做空')>=0)return 'row-short';return ''}
function buildTable(title,records,keyCols){
  if(!records||!records.length||!Array.isArray(records))return '<div class="table-wrap"><div class="table-title">'+title+'</div><div style="padding:20px;text-align:center;color:var(--text2)">暂无数据</div></div>';
  var allKeys=Object.keys(records[0]);
  var h='<div class="table-wrap"><div class="table-title">'+title+' ('+records.length+'条)</div><div style="overflow-x:auto"><table><thead><tr>';
  keyCols.forEach(function(c){if(allKeys.indexOf(c)>=0)h+='<th>'+c+'</th>'});
  h+='</tr></thead><tbody>';
  records.forEach(function(r){
    var adv=r['今日建议']||r['是否建议']||'';
    h+='<tr class="'+rowClass(adv)+'">';
    keyCols.forEach(function(c){
      if(allKeys.indexOf(c)<0)return;var v=r[c];
      if(c==='今日建议'||c==='方向偏好'||c==='历史回测偏好'||c==='数据新鲜度')v=badge(v);
      else if(c==='综合评分')v='<strong style="color:var(--blue)">'+fmt(v)+'</strong>';
      else if(c==='胜率'){var n=Number(v);v=n>0.55?'<span style="color:var(--green)">'+pct(n)+'</span>':n<0.45?'<span style="color:var(--red)">'+pct(n)+'</span>':pct(n)}
      else if(c==='平均收益'||c==='盈亏比'){var n=Number(v);v=n>0?'<span style="color:var(--green)">'+fmt(n)+'</span>':n<0?'<span style="color:var(--red)">'+fmt(n)+'</span>':fmt(n)}
      else if(c==='最大回撤'){var n=Number(v);v=n<-0.1?'<span style="color:var(--red)">'+fmt(n)+'</span>':fmt(n)}
      else if(c==='净分'){var n=Number(v);v=n>0?'<span style="color:var(--green)">'+v+'</span>':n<0?'<span style="color:var(--red)">'+v+'</span>':(v||'-')}
      else if(c==='ok')v=(v===true||v==='True')?'<span style="color:var(--green)">OK</span>':'<span style="color:var(--red)">FAIL</span>';
      else if(c==='实时快照状态')v=v==='已获取'?'<span style="color:var(--green)">已获取</span>':'<span style="color:var(--text2)">'+(v||'-')+'</span>';
      h+='<td title="'+(v!=null?v:'')+'">'+(v!=null?v:'-')+'</td>'});
    h+='</tr>'});
  h+='</tbody></table></div></div>';return h}
(function(){
  var D=DATA,R=D['今日看板'];
  if(!R||!Array.isArray(R)||!R.length){document.getElementById('panel-dashboard').innerHTML='<div style="text-align:center;padding:40px;color:var(--text2)">暂无看板数据</div>';return}
  var lC=0,sC=0,wC=0,aS=0;
  R.forEach(function(r){var a=r['今日建议']||'';if(a.indexOf('做多')>=0)lC++;else if(a.indexOf('做空')>=0)sC++;else wC++;aS+=Number(r['综合评分'])||0});
  aS=R.length?(aS/R.length).toFixed(1):0;
  var h='<div class="summary-grid"><div class="summary-card"><div class="label">总品种</div><div class="value" style="color:var(--blue)">'+R.length+'</div></div><div class="summary-card"><div class="label">建议做多</div><div class="value" style="color:var(--green)">'+lC+'</div></div><div class="summary-card"><div class="label">建议做空</div><div class="value" style="color:var(--red)">'+sC+'</div></div><div class="summary-card"><div class="label">观察/观望</div><div class="value" style="color:var(--orange)">'+wC+'</div></div><div class="summary-card"><div class="label">平均评分</div><div class="value" style="color:var(--blue)">'+aS+'</div></div></div>';
  h+='<div class="dash-cards">';
  R.forEach(function(r){
    var adv=r['今日建议']||'观望',cls=adv.indexOf('做多')>=0?'is-long':adv.indexOf('做空')>=0?'is-short':'is-watch',sc=adv.indexOf('做多')>=0?'var(--green)':adv.indexOf('做空')>=0?'var(--red)':'var(--orange)';
    h+='<div class="dash-card '+cls+'"><div class="dc-header"><div><span class="dc-symbol">'+(r['品种代码']||'')+'</span> <span class="dc-name">'+(r['品种名称']||'')+'</span></div><div class="dc-score" style="color:'+sc+'">'+fmt(r['综合评分'],1)+'</div></div><div style="margin-bottom:8px">'+badge(adv)+'</div><div style="font-size:11px;color:var(--text2);margin-bottom:6px">方向: '+(r['方向偏好']||'-')+' | 历史: '+(r['历史回测偏好']||'-')+'</div><div class="dc-metrics"><div class="dc-metric"><div class="dm-label">胜率</div><div class="dm-value">'+pct(Number(r['胜率']))+'</div></div><div class="dc-metric"><div class="dm-label">平均收益</div><div class="dm-value">'+fmt(r['平均收益'])+'</div></div><div class="dc-metric"><div class="dm-label">盈亏比</div><div class="dm-value">'+fmt(r['盈亏比'])+'</div></div><div class="dc-metric"><div class="dm-label">交易次数</div><div class="dm-value">'+(r['交易次数']||'-')+'</div></div><div class="dc-metric"><div class="dm-label">收盘价</div><div class="dm-value">'+(r['收盘价']||'-')+'</div></div><div class="dc-metric"><div class="dm-label">实时价</div><div class="dm-value">'+(r['实时价格']||'-')+'</div></div></div></div>'});
  h+='</div>';
  h+=buildTable('今日看板 - 明日交易看板',R,['品种代码','品种名称','今日建议','方向偏好','历史回测偏好','多因子数','空因子数','胜率','平均收益','盈亏比','最大回撤','交易次数','综合评分','收盘价','实时价格','近10日平均价格振幅','近10日最小价格振幅','最优参数']);
  document.getElementById('panel-dashboard').innerHTML=h})();
document.getElementById('panel-all').innerHTML=buildTable('全部活跃品种评分',DATA['全部品种评分'],['品种代码','品种名称','交易所','今日建议','多因子数','空因子数','胜率','平均收益','盈亏比','最大回撤','交易次数','综合评分','方向偏好','收盘价','成交量','持仓量','近10日平均价格振幅','近10日最小价格振幅','实时价格','盘中振幅率','最优参数']);
document.getElementById('panel-lowfreq').innerHTML=buildTable('高胜率低频机会池',DATA['高胜率低频'],['品种代码','品种名称','今日建议','方向偏好','历史回测偏好','胜率','平均收益','盈亏比','最大回撤','交易次数','综合评分','收盘价','实时价格','近10日平均价格振幅','最优参数']);
document.getElementById('panel-backtest').innerHTML=buildTable('回测参数与绩效明细',DATA['回测明细'],['品种代码','品种名称','参数','交易次数','胜率','平均收益','盈亏比','最大回撤','综合评分','多头次数','空头次数']);
(function(){
  var R=DATA['因子明细'],fc=['前高前低','大阳大阴','跳空缺口','颈线趋势线','整数关口','黄金分割','趋势方向','支撑阻力','成交量确认','持仓量确认','移动均线','形态突破','波动率变化'];
  var ak=R&&R.length?Object.keys(R[0]):[];
  document.getElementById('panel-factors').innerHTML=buildTable('最新交易日 13 因子明细',R,['品种代码','品种名称','多因子数','空因子数','净分'].concat(fc.filter(function(f){return ak.indexOf(f)>=0})))})();
(function(){
  var R=DATA['数据状态'],okC=0,fC=0;
  if(R&&Array.isArray(R))R.forEach(function(r){if(r.ok===true||r.ok==='True')okC++;else fC++});
  var h='<div class="summary-grid"><div class="summary-card"><div class="label">总品种</div><div class="value" style="color:var(--blue)">'+(R?R.length:0)+'</div></div><div class="summary-card"><div class="label">数据正常</div><div class="value" style="color:var(--green)">'+okC+'</div></div><div class="summary-card"><div class="label">数据异常</div><div class="value" style="color:var(--red)">'+fC+'</div></div></div>';
  h+=buildTable('数据更新状态',R,['symbol','name','exchange','ok','source','last_date','rows','数据新鲜度','avg_volume60','avg_hold60','实时快照状态']);
  document.getElementById('panel-status').innerHTML=h})();
document.getElementById('panel-realtime').innerHTML=buildTable('盘中实时快照',DATA['实时快照'],['品种代码','实时主力合约','行情时间','实时开盘','实时最高','实时最低','实时价格','买价','卖价','实时持仓','实时成交量','实时均价','昨收','盘中振幅','盘中振幅率']);
(function(){
  var R=DATA['参数说明'],h='<div class="table-wrap"><div class="table-title">参数说明</div><div style="overflow-x:auto"><table><thead><tr><th>参数</th><th>当前设置</th><th>说明</th></tr></thead><tbody>';
  if(R&&Array.isArray(R))R.forEach(function(r){h+='<tr><td>'+(r['参数']||'')+'</td><td>'+(r['当前设置']||'')+'</td><td>'+(r['说明']||'')+'</td></tr>'});
  h+='</tbody></table></div></div>';
  if(DATA.log)h+='<div class="table-wrap" style="margin-top:16px"><div class="table-title">运行日志</div><div class="log-box">'+DATA.log.replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;')+'</div></div>';
  document.getElementById('panel-params').innerHTML=h})();
</script>
</body>
</html>"""


def _export_static_html(
    out_dir: Path,
    stamp: str,
    dashboard_df: pd.DataFrame,
    all_df: pd.DataFrame,
    low_df: pd.DataFrame,
    backtest_df: pd.DataFrame,
    factor_df: pd.DataFrame,
    status_df: pd.DataFrame,
    realtime_df: pd.DataFrame,
    config: dict,
    logs: List[str],
    start_ts: datetime,
) -> Optional[Path]:
    """将所有数据内嵌到单个HTML文件中，可直接在手机浏览器打开。"""
    try:
        data = {
            "file_name": f"期货模型交易看板_{stamp}",
            "file_time": start_ts.strftime("%Y-%m-%d %H:%M:%S"),
            "今日看板": _df_to_records(dashboard_df) if not dashboard_df.empty else [],
            "全部品种评分": _df_to_records(all_df) if not all_df.empty else [],
            "高胜率低频": _df_to_records(low_df) if not low_df.empty else [],
            "回测明细": _df_to_records(backtest_df) if not backtest_df.empty else [],
            "因子明细": _df_to_records(factor_df) if not factor_df.empty else [],
            "数据状态": _df_to_records(status_df) if not status_df.empty else [],
            "实时快照": _df_to_records(realtime_df) if not realtime_df.empty else [],
            "参数说明": _df_to_records(pd.DataFrame(PARAM_EXPLAIN + FACTOR_EXPLAIN, columns=["参数", "当前设置", "说明"])),
            "log": "\n".join(logs),
        }
        json_str = json.dumps(data, ensure_ascii=False, default=str)
        html = _STATIC_HTML_TEMPLATE
        html = html.replace("__JSON_DATA__", json_str)
        html = html.replace("__FILE_NAME__", data["file_name"])
        html = html.replace("__FILE_TIME__", data["file_time"])
        html_path = out_dir / f"期货模型交易看板_{stamp}.html"
        html_path.write_text(html, encoding="utf-8")
        return html_path
    except Exception as e:
        print(f"[警告] 静态HTML生成失败（不影响Excel）: {e}")
        return None

def run(config: dict, mock: bool = False) -> Path:
    ensure_dirs(config)
    start_ts = datetime.now()
    stamp = start_ts.strftime("%Y%m%d_%H%M%S")
    out_dir = Path(config["output_dir"])
    log_path = out_dir / f"运行日志_{stamp}.txt"

    logs: List[str] = []

    def log(msg: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
        print(line)
        logs.append(line)

    log("读取品种列表...")
    symbols_df = fetch_symbol_list(config, mock=mock)
    log(f"待处理品种数：{len(symbols_df)}")

    data_map: Dict[str, pd.DataFrame] = {}
    status_rows: List[dict] = []
    meta = {}
    for row in tqdm(symbols_df.itertuples(index=False), total=len(symbols_df), desc="更新行情"):
        symbol = str(row.symbol).upper()
        exchange = str(row.exchange)
        name = clean_name(str(row.name))
        df, status = update_one_symbol(symbol, name, exchange, config, mock=mock)
        status_rows.append(status)
        if not df.empty:
            data_map[symbol] = df
            meta[symbol] = {"name": name, "exchange": exchange}

    if not data_map:
        raise RuntimeError("没有任何品种更新成功。请检查网络、akshare 版本或稍后重试。")

    # 流动性过滤：先算每个品种的平均成交量、持仓、最后日期
    liq_rows = []
    for symbol, df in data_map.items():
        tail = df.tail(60)
        liq_rows.append({
            "symbol": symbol,
            "avg_volume60": float(tail["volume"].mean()),
            "avg_hold60": float(tail["hold"].mean()),
            "last_date": df["date"].max(),
            "rows": len(df),
        })
    liq_df = pd.DataFrame(liq_rows)
    latest_global = liq_df["last_date"].max()
    global_freshness = diagnose_data_freshness(latest_global, start_ts)
    log(f"数据最新日期：{pd.Timestamp(latest_global).strftime('%Y-%m-%d') if pd.notna(latest_global) else '无'}；诊断：{global_freshness}")
    vol_cut = liq_df["avg_volume60"].quantile(float(config.get("liquidity_quantile", 0.25)))
    hold_cut = liq_df["avg_hold60"].quantile(float(config.get("liquidity_quantile", 0.25)))
    min_bars = int(config.get("min_bars", 180))
    max_stale = int(config.get("max_stale_calendar_days", 10))

    active_symbols = []
    for _, r in liq_df.iterrows():
        stale_days = (latest_global - r["last_date"]).days if pd.notna(latest_global) and pd.notna(r["last_date"]) else 999
        ok = (
            r["rows"] >= min_bars
            and r["avg_volume60"] >= vol_cut
            and r["avg_hold60"] >= hold_cut
            and stale_days <= max_stale
        )
        if ok:
            active_symbols.append(r["symbol"])
        # 更新状态行补充流动性
        for sr in status_rows:
            if sr["symbol"] == r["symbol"]:
                sr.update({
                    "avg_volume60": round(float(r["avg_volume60"]), 2),
                    "avg_hold60": round(float(r["avg_hold60"]), 2),
                    "stale_days": int(stale_days),
                    "liquidity_pass": bool(ok),
                    "运行时间": start_ts.strftime("%Y-%m-%d %H:%M:%S"),
                    "运行日": start_ts.strftime("%Y-%m-%d"),
                    "距运行日天数": int((pd.Timestamp(start_ts.date()) - pd.Timestamp(r["last_date"]).normalize()).days) if pd.notna(r["last_date"]) else "",
                    "数据新鲜度": diagnose_data_freshness(r["last_date"], start_ts),
                    "全市场最新日期": pd.Timestamp(latest_global).strftime("%Y-%m-%d") if pd.notna(latest_global) else "",
                    "全市场新鲜度诊断": global_freshness,
                })
    log(f"通过流动性过滤品种数：{len(active_symbols)}")

    # 合约规格：只对活跃品种获取，失败不阻断。
    log("获取合约规格/保证金信息（失败不影响模型运行）...")
    specs = fetch_contract_specs(active_symbols, mock=mock)

    log("尝试获取盘中实时快照（交易时段内用于当前价格/盘中信号，盘后不改变日线逻辑）...")
    live_map, realtime_df, realtime_msg = fetch_realtime_snapshots(active_symbols, data_map, config, mock=mock)
    log(realtime_msg)
    if not realtime_df.empty:
        realtime_df = realtime_df.copy()
        # 仅保留更适合阅读的列名，实际计算仍使用 live_map。
        realtime_df["盘中振幅"] = realtime_df["high"] - realtime_df["low"]
        realtime_df["盘中振幅率"] = realtime_df["盘中振幅"] / realtime_df["last_settle_price"].replace(0, np.nan)
        realtime_df.rename(columns={
            "symbol": "品种代码",
            "contract": "实时主力合约",
            "spot_symbol": "接口品种名",
            "spot_time": "行情时间",
            "open": "实时开盘",
            "high": "实时最高",
            "low": "实时最低",
            "current_price": "实时价格",
            "bid_price": "买价",
            "ask_price": "卖价",
            "hold": "实时持仓",
            "volume": "实时成交量",
            "avg_price": "实时均价",
            "last_close": "昨收",
            "last_settle_price": "昨结算",
            "source": "实时数据源",
        }, inplace=True)
    for sr in status_rows:
        sym = sr.get("symbol")
        sr["实时快照状态"] = "已获取" if sym in live_map else "未获取/盘后可忽略"
        sr["实时快照说明"] = realtime_msg

    all_results: List[dict] = []
    lowfreq_rows: List[dict] = []
    factor_latest_rows: List[dict] = []
    backtest_rows: List[dict] = []

    for symbol in tqdm(active_symbols, desc="计算因子与回测"):
        raw = data_map[symbol]
        name = meta[symbol]["name"]
        exchange = meta[symbol]["exchange"]
        complete_daily_date = raw["date"].max() if not raw.empty else pd.NaT
        complete_daily_close = safe_float(raw.iloc[-1].get("close")) if not raw.empty else np.nan
        range_stats = calc_range_stats_from_daily(raw, 10)
        live = live_map.get(symbol)
        fac_input, market_mode = append_intraday_snapshot(raw, live, start_ts, config)
        fac = calculate_factors(fac_input)
        if fac.empty:
            continue
        # 只保留最近 lookback_bars+1 做回测与报告；若盘中有实时快照，该快照仅参与最新信号，不参与有结果的历史回测。
        fac = fac.tail(int(config.get("lookback_bars", 250)) + 1).reset_index(drop=True)
        best_regular, best_lowfreq, bt_all = find_best_models(symbol, name, fac, config)

        latest = fac.iloc[-1]
        spec = specs.get(symbol, {})
        multiplier = safe_float(spec.get("multiplier", np.nan))
        margin_ratio = safe_float(spec.get("margin_ratio", np.nan))
        close = safe_float(latest.get("close"))
        settle = safe_float(latest.get("settle"))
        margin_amount = close * multiplier * margin_ratio if not any(pd.isna(x) for x in [close, multiplier, margin_ratio]) else np.nan

        advice, sig, bull, bear = latest_signal(fac, best_regular)
        current_bias = current_bias_text(advice, sig, bull, bear)
        signal_strength = abs(bull - bear)
        is_reco = 1 if advice in ["建议做多", "建议做空"] else 0

        if best_regular is not None:
            all_results.append({
                "日期": latest["date"].strftime("%Y-%m-%d"),
                "品种代码": symbol,
                "品种名称": name,
                "交易所": exchange,
                "数据口径": "品种指数/连续合约",
                "当前合约": spec.get("main_contract", symbol),
                "今日建议": advice,
                "是否建议": is_reco,
                "信号强度": signal_strength,
                "多因子数": bull,
                "空因子数": bear,
                "最优参数": f"至少{best_regular.min_factors}因子，净差≥{best_regular.net_gap}",
                "胜率": best_regular.win_rate,
                "平均收益": best_regular.avg_return,
                "盈亏比": best_regular.profit_factor,
                "最大回撤": best_regular.max_drawdown,
                "交易次数": best_regular.trades,
                "综合评分": best_regular.composite_score,
                "方向偏好": current_bias,
                "历史回测偏好": best_regular.direction_bias,
                "收盘价": close,
                "结算价": settle,
                "成交量": int(latest.get("volume", 0)),
                "持仓量": int(latest.get("hold", 0)),
                "保证金比例": margin_ratio,
                "合约乘数": multiplier,
                "估算保证金/手": margin_amount,
                "规格备注": spec.get("spec_message", ""),
            })
        else:
            # 没有满足常规交易次数，也纳入观察。
            all_results.append({
                "日期": latest["date"].strftime("%Y-%m-%d"),
                "品种代码": symbol,
                "品种名称": name,
                "交易所": exchange,
                "数据口径": "品种指数/连续合约",
                "当前合约": spec.get("main_contract", symbol),
                "今日建议": advice,
                "是否建议": 0,
                "信号强度": signal_strength,
                "多因子数": bull,
                "空因子数": bear,
                "最优参数": "无常规模型",
                "胜率": np.nan,
                "平均收益": np.nan,
                "盈亏比": np.nan,
                "最大回撤": np.nan,
                "交易次数": 0,
                "综合评分": 0,
                "方向偏好": current_bias,
                "历史回测偏好": "",
                "收盘价": close,
                "结算价": settle,
                "成交量": int(latest.get("volume", 0)),
                "持仓量": int(latest.get("hold", 0)),
                "保证金比例": margin_ratio,
                "合约乘数": multiplier,
                "估算保证金/手": margin_amount,
                "规格备注": spec.get("spec_message", ""),
            })

        extra_fields = {
            "行情模式": market_mode,
            "完整日线日期": pd.Timestamp(complete_daily_date).strftime("%Y-%m-%d") if pd.notna(complete_daily_date) else "",
            "完整日线收盘价": complete_daily_close,
            "近10日平均价格振幅": range_stats.get("avg_range10", np.nan),
            "近10日最小价格振幅": range_stats.get("min_range10", np.nan),
            "近10日最大价格振幅": range_stats.get("max_range10", np.nan),
            "近10日平均振幅率": range_stats.get("avg_range10_pct", np.nan),
            "近10日最小振幅率": range_stats.get("min_range10_pct", np.nan),
        }
        extra_fields.update(live_fields(live, market_mode))
        if all_results:
            all_results[-1].update(extra_fields)

        if best_lowfreq is not None:
            low_advice, low_sig, lbull, lbear = latest_signal(fac, best_lowfreq)
            low_current_bias = current_bias_text(low_advice, low_sig, lbull, lbear)
            lowfreq_rows.append({
                "日期": latest["date"].strftime("%Y-%m-%d"),
                "品种代码": symbol,
                "品种名称": name,
                "今日建议": low_advice,
                "方向偏好": low_current_bias,
                "历史回测偏好": best_lowfreq.direction_bias,
                "今日信号分": abs(lbull - lbear),
                "最优参数": f"至少{best_lowfreq.min_factors}因子，净差≥{best_lowfreq.net_gap}",
                "胜率": best_lowfreq.win_rate,
                "平均收益": best_lowfreq.avg_return,
                "盈亏比": best_lowfreq.profit_factor,
                "最大回撤": best_lowfreq.max_drawdown,
                "交易次数": best_lowfreq.trades,
                "综合评分": best_lowfreq.composite_score,
                "收盘价": close,
                "结算价": settle,
            })
            lowfreq_rows[-1].update(extra_fields)

        if not bt_all.empty:
            bt_top = bt_all.sort_values("composite_score", ascending=False).head(5).copy()
            for _, b in bt_top.iterrows():
                backtest_rows.append({
                    "品种代码": symbol,
                    "品种名称": name,
                    "参数": f"至少{int(b['min_factors'])}因子，净差≥{int(b['net_gap'])}",
                    "交易次数": int(b["trades"]),
                    "胜率": float(b["win_rate"]),
                    "平均收益": float(b["avg_return"]),
                    "盈亏比": float(b["profit_factor"]),
                    "最大回撤": float(b["max_drawdown"]),
                    "综合评分": float(b["composite_score"]),
                    "多头次数": int(b.get("long_trades", 0)),
                    "空头次数": int(b.get("short_trades", 0)),
                })

        frow = {
            "日期": latest["date"].strftime("%Y-%m-%d"),
            "品种代码": symbol,
            "品种名称": name,
            "多因子数": bull,
            "空因子数": bear,
            "净分": int(latest.get("net_factor_score", 0)),
        }
        frow.update({
            "行情模式": market_mode,
            "完整日线日期": pd.Timestamp(complete_daily_date).strftime("%Y-%m-%d") if pd.notna(complete_daily_date) else "",
            "近10日平均价格振幅": range_stats.get("avg_range10", np.nan),
            "近10日最小价格振幅": range_stats.get("min_range10", np.nan),
            "实时价格": live_fields(live, market_mode).get("实时价格", np.nan),
        })
        for fname in FACTOR_NAMES:
            val = int(latest.get(f"因子_{fname}", 0))
            frow[fname] = sign_to_text(val)
        factor_latest_rows.append(frow)

    all_df, low_df, status_df, factor_df = build_report_rows(all_results, lowfreq_rows, status_rows, factor_latest_rows)
    backtest_df = pd.DataFrame(backtest_rows)
    if not backtest_df.empty:
        backtest_df = backtest_df.sort_values(["品种代码", "综合评分"], ascending=[True, False])

    # 今日看板：优先建议做多/做空；如果不足 top_n，则用观察偏多/偏空补足。
    top_n = int(config.get("top_n", 5))
    if all_df.empty:
        dashboard_df = pd.DataFrame()
    else:
        reco = all_df[all_df["今日建议"].isin(["建议做多", "建议做空"])].sort_values(
            ["综合评分", "信号强度"], ascending=[False, False]
        )
        obs = all_df[~all_df.index.isin(reco.index)].sort_values(["信号强度", "综合评分"], ascending=[False, False])
        dashboard_df = pd.concat([reco, obs], ignore_index=True).head(top_n)
        keep_cols = [
            "日期", "品种代码", "品种名称", "行情模式", "当前合约", "实时合约", "今日建议", "方向偏好", "历史回测偏好",
            "多因子数", "空因子数", "胜率", "平均收益", "盈亏比", "最大回撤", "交易次数", "综合评分",
            "收盘价", "结算价", "实时价格", "实时最高", "实时最低", "盘中振幅", "盘中振幅率",
            "近10日平均价格振幅", "近10日最小价格振幅", "完整日线日期", "完整日线收盘价",
            "保证金比例", "估算保证金/手", "最优参数"
        ]
        dashboard_df = dashboard_df[[c for c in keep_cols if c in dashboard_df.columns]]

    xlsx_path = out_dir / f"期货模型交易看板_{stamp}.xlsx"
    log(f"输出 Excel：{xlsx_path}")
    export_excel(xlsx_path, dashboard_df, all_df, low_df, backtest_df, factor_df, status_df, realtime_df, config)

    # 自动生成静态HTML（数据内嵌，手机直接打开）
    html_path = _export_static_html(
        out_dir, stamp, dashboard_df, all_df, low_df, backtest_df,
        factor_df, status_df, realtime_df, config, logs, start_ts,
    )

    logs.append(f"输出文件：{xlsx_path}")
    if html_path:
        logs.append(f"静态HTML：{html_path}")
    logs.append(f"开始时间：{start_ts}")
    logs.append(f"结束时间：{datetime.now()}")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(logs))
    log(f"完成：{xlsx_path}")
    if html_path:
        log(f"静态HTML：{html_path}（手机直接打开即可查看）")
    return xlsx_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config.json", help="配置文件路径")
    parser.add_argument("--mock", action="store_true", help="使用模拟数据测试程序，不联网")
    args = parser.parse_args()

    try:
        config = load_config(args.config)
        mock = bool(args.mock or config.get("mock_mode", False))
        path = run(config, mock=mock)
        print("\n===================== 运行成功 =====================")
        print(f"结果文件：{path}")
        print("请打开 output 文件夹查看 Excel。")
        print("====================================================")
        return 0
    except Exception as e:
        Path("output").mkdir(exist_ok=True)
        err_path = Path("output") / f"错误日志_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(err_path, "w", encoding="utf-8") as f:
            f.write(str(e) + "\n\n")
            f.write(traceback.format_exc())
        print("\n===================== 运行失败 =====================")
        print(str(e))
        print(f"错误日志：{err_path}")
        print("常见处理：1）检查网络；2）升级 akshare；3）先运行 RUN_MOCK_TEST.bat 验证本地环境。")
        print("====================================================")
        return 1


if __name__ == "__main__":
    sys.exit(main())
